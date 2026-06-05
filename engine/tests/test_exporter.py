from decimal import Decimal
from unittest.mock import patch

import openpyxl
import polars as pl
import pytest

from engine.export_config import ExportConfig, Output
from engine.exporter import export_outputs


def _xlsx(filename, sheets):
    return Output(filename=filename, format="xlsx", sheets=sheets)


def _csv(filename, query):
    return Output(filename=filename, format="csv", query=query)


def test_xlsx_multi_sheet_from_multiple_queries(tmp_path):
    results = {
        "q_alpha": pl.DataFrame({"v": [1, 2]}),
        "q_beta": pl.DataFrame({"v": [3, 4]}),
    }
    config = ExportConfig(
        outputs=[_xlsx("report.xlsx", {"Alpha": "q_alpha", "Beta": "q_beta"})]
    )
    outcomes = export_outputs(results, config, tmp_path)

    assert outcomes["report.xlsx"].status == "written"
    wb = openpyxl.load_workbook(tmp_path / "report.xlsx")
    assert wb.sheetnames == ["Alpha", "Beta"]


def test_csv_output_written(tmp_path):
    results = {"q": pl.DataFrame({"a": [1, 2], "b": ["x", "y"]})}
    config = ExportConfig(outputs=[_csv("out.csv", "q")])
    outcomes = export_outputs(results, config, tmp_path)

    assert outcomes["out.csv"].status == "written"
    content = (tmp_path / "out.csv").read_text(encoding="utf-8")
    assert content == "a,b\n1,x\n2,y\n"


def test_missing_query_marks_output_failed(tmp_path):
    results = {"present": pl.DataFrame({"v": [1]})}
    config = ExportConfig(outputs=[_csv("out.csv", "absent")])
    outcomes = export_outputs(results, config, tmp_path)

    assert outcomes["out.csv"].status == "failed"
    assert "absent" in outcomes["out.csv"].reason
    assert not (tmp_path / "out.csv").exists()


def test_output_folder_created_if_missing(tmp_path):
    deep = tmp_path / "a" / "b" / "c"
    results = {"q": pl.DataFrame({"v": [1]})}
    config = ExportConfig(outputs=[_csv("out.csv", "q")])
    export_outputs(results, config, deep)
    assert (deep / "out.csv").exists()


def test_no_tmp_files_remain(tmp_path):
    results = {"q": pl.DataFrame({"v": [1, 2, 3]})}
    config = ExportConfig(
        outputs=[_csv("a.csv", "q"), _xlsx("b.xlsx", {"S": "q"})]
    )
    export_outputs(results, config, tmp_path)
    assert list(tmp_path.glob("*.tmp-*")) == []


def test_decimal_precision_preserved_in_xlsx(tmp_path):
    df = pl.DataFrame({"amount": ["0.10"] * 1000}).with_columns(
        pl.col("amount").cast(pl.Decimal(scale=2))
    )
    config = ExportConfig(outputs=[_xlsx("dec.xlsx", {"S": "q"})])
    export_outputs({"q": df}, config, tmp_path)

    wb = openpyxl.load_workbook(tmp_path / "dec.xlsx", data_only=True)
    ws = wb.active
    read_sum = Decimal("0")
    sample_cell = None
    for row in ws.iter_rows(min_row=2):
        cell = row[0]
        if sample_cell is None:
            sample_cell = cell
        read_sum += Decimal(str(cell.value))
    assert read_sum == Decimal("100.00")
    assert sample_cell.number_format == "#,##0.00"
    assert sample_cell.data_type == "n"


def test_decimal_precision_preserved_in_csv(tmp_path):
    # 1000 x 0.10 written as exact strings, summed back -> exactly 100.00.
    df = pl.DataFrame({"amount": ["0.10"] * 1000}).with_columns(
        pl.col("amount").cast(pl.Decimal(scale=2))
    )
    config = ExportConfig(outputs=[_csv("dec.csv", "q")])
    export_outputs({"q": df}, config, tmp_path)

    lines = (tmp_path / "dec.csv").read_text(encoding="utf-8").splitlines()
    assert lines[0] == "amount"
    values = [Decimal(v) for v in lines[1:]]
    assert all(v == Decimal("0.10") for v in values)
    assert sum(values) == Decimal("100.00")


def test_decimal_three_places_format_in_xlsx(tmp_path):
    df = pl.DataFrame({"price": ["1.234", "5.678"]}).with_columns(
        pl.col("price").cast(pl.Decimal(scale=3))
    )
    config = ExportConfig(outputs=[_xlsx("p.xlsx", {"P": "q"})])
    export_outputs({"q": df}, config, tmp_path)

    wb = openpyxl.load_workbook(tmp_path / "p.xlsx", data_only=True)
    cell = list(wb.active.iter_rows(min_row=2))[0][0]
    assert cell.number_format == "#,##0.000"


def test_rename_failure_marks_output_failed_no_partial(tmp_path):
    results = {"q": pl.DataFrame({"v": [1]})}
    config = ExportConfig(outputs=[_csv("locked.csv", "q")])
    with patch("pathlib.Path.replace", side_effect=PermissionError("file in use")):
        outcomes = export_outputs(results, config, tmp_path)
    assert outcomes["locked.csv"].status == "failed"
    assert "PermissionError" in outcomes["locked.csv"].reason
    assert not (tmp_path / "locked.csv").exists()
    assert list(tmp_path.glob("*.tmp-*")) == []
