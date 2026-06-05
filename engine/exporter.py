"""Write output files from in-memory query results, driven by `exports.json`.

One query produces one table; an output file is assembled from one or more of
those tables (an xlsx maps several sheets to several queries; a csv is one
query). Writes are atomic (temp file then rename) so a target left open in
Power BI fails loudly on the rename rather than leaving a half-written file.

Decimal precision: CSV is written with Decimal columns rendered as their exact
string form (never through float). Excel has no decimal type, so the xlsx path
necessarily stores float64 with a Decimal-aware display format — exact for the
~15 significant digits finance figures occupy, but not arbitrary-precision.
"""

import uuid
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import polars as pl

from engine.export_config import ExportConfig, Output, _validate_sheet_name


@dataclass
class OutputResult:
    status: str  # "written" | "failed"
    reason: str | None = None


def _decimal_places_for_column(col: pl.Series) -> int:
    dtype = col.dtype
    if isinstance(dtype, pl.Decimal):
        scale = dtype.scale
        return scale if scale is not None else 2
    return 2


def _is_decimal_column(col: pl.Series) -> bool:
    return isinstance(col.dtype, pl.Decimal)


def _atomic_write(target: Path, write_tmp) -> None:
    """Write to a temp file via `write_tmp(tmp_path)`, then rename onto target.

    The rename is what fails (loudly) if the target is locked by Power BI.
    """
    target.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = target.parent / f"{target.name}.tmp-{uuid.uuid4()}{target.suffix}"
    try:
        write_tmp(tmp_path)
        tmp_path.replace(target)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def _write_xlsx(sheets: dict[str, pl.DataFrame], target: Path) -> None:
    for name in sheets:
        _validate_sheet_name(name)

    def _build(tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name, df in sheets.items():
            ws = wb.create_sheet(title=sheet_name)

            decimal_columns: dict[int, str] = {}
            for col_idx, col_name in enumerate(df.columns):
                if _is_decimal_column(df[col_name]):
                    places = _decimal_places_for_column(df[col_name])
                    decimal_columns[col_idx] = "#,##0." + "0" * places

            ws.append(list(df.columns))

            # Cast Decimal columns to Float64 for openpyxl (needs numeric scalars).
            df_export = df
            for col_idx, col_name in enumerate(df.columns):
                if col_idx in decimal_columns:
                    df_export = df_export.with_columns(pl.col(col_name).cast(pl.Float64))

            for row_tuple in df_export.iter_rows():
                ws.append(list(row_tuple))

            if decimal_columns:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for col_idx, fmt in decimal_columns.items():
                        row[col_idx].number_format = fmt

        wb.save(tmp_path)

    _atomic_write(target, _build)


def _write_csv(df: pl.DataFrame, target: Path) -> None:
    # Render Decimal columns as exact strings so no float widening occurs.
    df_export = df
    for col_name in df.columns:
        if _is_decimal_column(df[col_name]):
            df_export = df_export.with_columns(pl.col(col_name).cast(pl.String))

    _atomic_write(target, lambda tmp_path: df_export.write_csv(tmp_path))


def export_outputs(
    results: dict[str, pl.DataFrame],
    config: ExportConfig,
    output_folder: str | Path,
) -> dict[str, OutputResult]:
    """Write every output in `config` from the in-memory result cache.

    An output whose query is absent from `results` (because it failed or was
    skipped) is not written and is marked failed. Each output is independent —
    one failure does not stop the others.
    """
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    outcomes: dict[str, OutputResult] = {}
    for out in config.outputs:
        missing = sorted(out.query_names() - results.keys())
        if missing:
            outcomes[out.filename] = OutputResult(
                "failed",
                f"required quer{'y' if len(missing) == 1 else 'ies'} "
                f"{missing} did not produce output",
            )
            continue

        target = output_folder / out.filename
        try:
            if out.format == "xlsx":
                sheets = {sheet: results[q] for sheet, q in out.sheets.items()}
                _write_xlsx(sheets, target)
            else:
                _write_csv(results[out.query], target)
            outcomes[out.filename] = OutputResult("written")
        except Exception as e:
            outcomes[out.filename] = OutputResult("failed", f"{type(e).__name__}: {e}")

    return outcomes
