import json
import os
from pathlib import Path

import openpyxl
import polars as pl
from openpyxl.utils import range_boundaries

_project_root: Path | None = None


def _find_project_root() -> Path:
    global _project_root
    if _project_root is not None:
        return _project_root
    candidate = Path(__file__).resolve().parent
    while candidate != candidate.parent:
        if (candidate / "pyproject.toml").exists():
            _project_root = candidate
            return _project_root
        candidate = candidate.parent
    raise FileNotFoundError("pyproject.toml not found walking up from engine/loader.py")


def _load_config() -> dict:
    env_path = os.environ.get("PIPELINE_CONFIG")
    if env_path:
        config_path = Path(env_path)
    else:
        config_path = _find_project_root() / "config.json"
    with config_path.open(encoding="utf-8") as f:
        return json.load(f)


def _read_csv_path(path: str | Path) -> pl.DataFrame:
    return pl.read_csv(path)


def _read_excel_path(path: str | Path, sheet_or_table: str) -> pl.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)

    if sheet_or_table in wb.sheetnames:
        ws = wb[sheet_or_table]
        rows = list(ws.values)
        if not rows:
            return pl.DataFrame()
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data = {h: [row[i] for row in rows[1:]] for i, h in enumerate(headers)}
        return pl.DataFrame(data)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for table_name, ref in ws.tables.items():
            if table_name == sheet_or_table:
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                rows = list(
                    ws.iter_rows(
                        min_row=min_row,
                        max_row=max_row,
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True,
                    )
                )
                if not rows:
                    return pl.DataFrame()
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
                data = {h: [row[i] for row in rows[1:]] for i, h in enumerate(headers)}
                return pl.DataFrame(data)

    raise KeyError(
        f"'{sheet_or_table}' is not a sheet name or table name in '{path}'"
    )


def read_excel(source_name: str, sheet_or_table: str) -> pl.DataFrame:
    config = _load_config()
    section = "excel_sources"
    if source_name not in config.get(section, {}):
        raise KeyError(
            f"'{source_name}' not found in '{section}' in config.json"
        )
    path = config[section][source_name]
    return _read_excel_path(path, sheet_or_table)


def _read_jsonl_path(path: str | Path) -> pl.DataFrame:
    records = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    if not records:
        return pl.DataFrame()
    return pl.DataFrame(records)


def read_jsonl(source_name: str) -> pl.DataFrame:
    config = _load_config()
    section = "jsonl_sources"
    if source_name not in config.get(section, {}):
        raise KeyError(
            f"'{source_name}' not found in '{section}' in config.json"
        )
    path = config[section][source_name]
    return _read_jsonl_path(path)


def read_csv(source_name: str) -> pl.DataFrame:
    config = _load_config()
    section = "csv_sources"
    if source_name not in config.get(section, {}):
        raise KeyError(
            f"'{source_name}' not found in '{section}' in config.json"
        )
    path = config[section][source_name]
    return _read_csv_path(path)


def read_sql(connection_name: str, query: str) -> pl.DataFrame:
    import pyodbc

    config = _load_config()
    section = "connections"
    if connection_name not in config.get(section, {}):
        raise KeyError(
            f"'{connection_name}' not found in '{section}' in config.json"
        )
    conn_cfg = config[section][connection_name]
    conn_type = conn_cfg.get("type")
    if conn_type == "mssql":
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={conn_cfg['host']};"
            f"DATABASE={conn_cfg['database']};"
            f"UID={conn_cfg['username']};"
            f"PWD={conn_cfg['password']}"
        )
        conn = pyodbc.connect(conn_str)
        try:
            cursor = conn.cursor()
            cursor.execute(query)
            columns = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()
            data = {col: [row[i] for row in rows] for i, col in enumerate(columns)}
            return pl.DataFrame(data)
        finally:
            conn.close()
    else:
        raise ValueError(f"Unsupported connection type: '{conn_type}'")
